import openpyxl
from openpyxl.utils import get_column_letter
import json

wb = openpyxl.load_workbook(
    r'c:\Users\kseaman\OneDrive - UAGC\General - Performance Marketing\Performance Marketing\Reporting Templates\MonthlyReporting\2026 - 04\Channel CPE Master - April.xlsx',
    data_only=True
)

print(f'Total sheets: {len(wb.sheetnames)}')
print(f'Sheet names: {wb.sheetnames}')
print()

for sheet_name in wb.sheetnames:
    ws = wb[sheet_name]
    print(f'=== Sheet: "{sheet_name}" ({ws.max_row} rows x {ws.max_column} cols) ===')
    
    # Print header row(s) - first 2 rows
    for row_idx in range(1, min(4, ws.max_row + 1)):
        row_data = []
        for col_idx in range(1, min(ws.max_column + 1, 20)):
            cell = ws.cell(row=row_idx, column=col_idx)
            val = cell.value
            if val is not None:
                row_data.append(f'{get_column_letter(col_idx)}: {str(val)[:40]}')
        if row_data:
            print(f'  Row {row_idx}: {row_data}')
    
    # Check for charts
    if ws._charts:
        print(f'  CHARTS: {len(ws._charts)} chart(s)')
        for idx, chart in enumerate(ws._charts):
            print(f'    Chart {idx+1}: type={chart.__class__.__name__}, title={chart.title}')
    
    # Sample a few more rows to understand data structure
    if ws.max_row > 5:
        print(f'  ... (showing rows 4-8 sample)')
        for row_idx in range(4, min(9, ws.max_row + 1)):
            row_data = []
            for col_idx in range(1, min(ws.max_column + 1, 15)):
                cell = ws.cell(row=row_idx, column=col_idx)
                val = cell.value
                if val is not None:
                    row_data.append(f'{get_column_letter(col_idx)}: {str(val)[:35]}')
            if row_data:
                print(f'    Row {row_idx}: {row_data}')
    
    print()
