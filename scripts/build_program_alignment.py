"""
Build program_alignment.json from the academic alignment Excel workbook.

Usage (from repo root):
  uvx --with openpyxl python scripts/build_program_alignment.py
  uvx --with openpyxl python scripts/build_program_alignment.py --xlsx path/to/file.xlsx
"""
from __future__ import annotations

import argparse
import json
from datetime import date
from pathlib import Path

ROOT = Path(__file__).resolve().parents[1]
DEFAULT_XLSX = ROOT / "data" / "program_alignment_source.xlsx"
OUT_PATH = ROOT / "data" / "program_alignment.json"

# Report program_name -> Excel Program Name (when they differ)
PROGRAM_NAME_ALIASES: dict[str, str] = {
    "Post Baccalaureate Teaching Certificate - Elementary Education": (
        "Post- Baccalaureate Teaching Certification"
    ),
}

FIELD_MAP = {
    "Program Name": "program_name",
    "Catalog Program Name": "catalog_program_name",
    "APL Program ID": "apl_program_id",
    "Active": "active",
    "Status": "status",
    "College": "college",
    "APL College ID": "apl_college_id",
    "Division": "division",
    "APL Division ID": "apl_division_id",
    "Department": "department",
    "APL Department ID": "apl_department_id",
    "Dean": "dean",
    "Associate Dean": "associate_dean",
    "Department Head": "department_head",
}


def _clean(value: object) -> str | None:
    if value is None:
        return None
    text = str(value).strip()
    return text or None


def _row_record(headers: tuple, row: tuple) -> dict:
    raw = dict(zip(headers, row))
    rec: dict[str, str | None] = {}
    for src, dest in FIELD_MAP.items():
        val = _clean(raw.get(src))
        if dest == "apl_program_id" and val:
            val = val.upper()
        rec[dest] = val
    return rec


def load_workbook_rows(xlsx_path: Path) -> list[dict]:
    import openpyxl

    wb = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]
    rows_iter = ws.iter_rows(values_only=True)
    headers = tuple(next(rows_iter))
    records: list[dict] = []
    for row in rows_iter:
        if not row or not row[0]:
            continue
        records.append(_row_record(headers, row))
    wb.close()
    return records


def build_indexes(records: list[dict]) -> tuple[dict, dict]:
    """Prefer Active=Yes row per program name; APL id indexes last active row."""
    by_name: dict[str, dict] = {}
    by_apl: dict[str, dict] = {}

    def score(rec: dict) -> int:
        active = (rec.get("active") or "").lower() == "yes"
        return 2 if active else 1

    for rec in records:
        name = rec.get("program_name")
        if name:
            prev = by_name.get(name)
            if prev is None or score(rec) >= score(prev):
                by_name[name] = rec
        apl = rec.get("apl_program_id")
        if apl:
            prev = by_apl.get(apl)
            if prev is None or score(rec) >= score(prev):
                by_apl[apl] = rec
    return by_name, by_apl


def main() -> None:
    parser = argparse.ArgumentParser(description="Build program_alignment.json")
    parser.add_argument(
        "--xlsx",
        type=Path,
        default=DEFAULT_XLSX,
        help="Path to Program Alignment List workbook",
    )
    args = parser.parse_args()
    xlsx_path = args.xlsx.resolve()
    if not xlsx_path.is_file():
        raise SystemExit(f"Workbook not found: {xlsx_path}")

    records = load_workbook_rows(xlsx_path)
    by_name, by_apl = build_indexes(records)

    payload = {
        "generated": str(date.today()),
        "source_file": xlsx_path.name,
        "row_count": len(records),
        "active_program_names": sum(
            1 for r in by_name.values() if (r.get("active") or "").lower() == "yes"
        ),
        "aliases": PROGRAM_NAME_ALIASES,
        "by_program_name": by_name,
        "by_apl_program_id": by_apl,
    }
    OUT_PATH.parent.mkdir(parents=True, exist_ok=True)
    OUT_PATH.write_text(json.dumps(payload, indent=2), encoding="utf-8")
    print(f"Saved {OUT_PATH}")
    print(f"  Rows read: {len(records)}")
    print(f"  By program name: {len(by_name)}")
    print(f"  By APL ID: {len(by_apl)}")


if __name__ == "__main__":
    main()
